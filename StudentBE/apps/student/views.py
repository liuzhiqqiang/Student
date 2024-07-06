import os.path

import openpyxl
from django.shortcuts import render
# 引入Student类
from student.models import Student
# 引入JsonResponse模块
from django.http import JsonResponse
# 导入json模块
import json
# 导入Q查询
from django.db.models import Q
# 导入uuid类
import uuid
# 导入哈希库
import hashlib
# 导入settings
from django.conf import settings

# Create your views here.
def get_students(request):
    # 获取所有学生的信息

    try:
        # 使用ORM获取所有学生信息
        obj_students = Student.objects.all().values()
        # 把结果转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, "data": students})
    except Exception as e:
        #如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取学生信息异常，具体错误：" + str(e)})

def query_students(request):
    # 查询学生信息
    # 获取所有学生的信息
    # 接口传递过来的查询条件 Axios默认是json格式---字典类型（‘inputstr’） --data['inputsre']
    data = json.loads(request.body.decode('utf-8'))

    try:
        # 使用ORM获取满足条件的学生信息
        obj_students = Student.objects.filter(Q(sno__icontains=data['inputstr']) | Q(name__icontains=data['inputstr']) |
            Q(gender__icontains=data['inputstr']) | Q(mobile__icontains=data['inputstr'])|
            Q(email__icontains=data['inputstr']) | Q(address__icontains=data['inputstr'])).values()

        # 把结果转为List
        students = list(obj_students)

        # 返回
        return JsonResponse({'code': 1, "data": students})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "查询学生信息异常，具体错误：" + str(e)})

def check_sno(request):
    # 判断要添加的学号是否存在
    data = json.loads(request.body.decode('utf-8'))

    try:
        obj_student = Student.objects.filter(sno=data['sno'])
        if obj_student.count() != 0:
            return JsonResponse({'code': 1, 'exists': True})
        else:
            return JsonResponse({'code': 1, 'exists': False})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg':"校验学号失败，具体原因：" + str(e)})

def add_student(request):
    # 添加学生到数据库
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    # 添加到数据库
    try:
        obj_student = Student(sno=data['sno'], name=data['name'],gender=data['gender'],
                              birthday=data['birthday'], mobile=data['mobile'],
                              email=data['email'], address=data['address'],
                              image=data['image'])
        # 执行添加
        obj_student.save()

        # 肯定是要更新页面中的学生内容的。
        # 使用ORM获取所有学生信息
        obj_students = Student.objects.all().values()
        # 把结果转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, "data": students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '添加到数据库出现异常，具体原因：' + str(e)})

def update_student(request):
    # 修改学生到数据库
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    # 添加到数据库
    try:
        # 先把原来的删除掉，然后直接复制添加的部分是可以的，当然也可以一个一个修改。
        # 先找到然后直接删除
        obj_student = Student.objects.get(sno=data['sno'])
        obj_student.delete()
        # 然后添加
        obj_student = Student(sno=data['sno'], name=data['name'],gender=data['gender'],
                              birthday=data['birthday'], mobile=data['mobile'],
                              email=data['email'], address=data['address'],
                              image=data['image'])
        # 执行添加
        obj_student.save()

        # 肯定是要更新页面中的学生内容的。
        # 使用ORM获取所有学生信息
        obj_students = Student.objects.all().values()
        # 把结果转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, "data": students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '修改数据库出现异常，具体原因：' + str(e)})

def delete_student(request):
    # 删除学生数据

    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    # 添加到数据库
    try:
        # 先找到然后直接删除
        obj_student = Student.objects.get(sno=data['sno'])
        obj_student.delete()

        # 肯定是要更新页面中的学生内容的。
        # 使用ORM获取所有学生信息
        obj_students = Student.objects.all().values()
        # 把结果转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, "data": students})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '删除数据库出现异常，具体原因：' + str(e)})

def delete_students(request):
    # 接收前端传递过来的值
    data = json.loads(request.body.decode('utf-8'))
    # 添加到数据库
    try:
        # 变量每一个学生并删除
        for stu in data['students']:
            obj_student = Student.objects.get(sno=stu['sno'])
            obj_student.delete()

        # 肯定是要更新页面中的学生内容的。
        # 使用ORM获取所有学生信息
        obj_students = Student.objects.all().values()
        # 把结果转为List
        students = list(obj_students)
        # 返回
        return JsonResponse({'code': 1, "data": students})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': '批量删除数据库出现异常，具体原因：' + str(e)})

# 获取图片的唯一名字
def get_unique_id():
    # 获取uuid的随机数
    uuid_val = uuid.uuid4()

    # 获取uuid的随机数字符串
    uuid_str = str(uuid_val).encode('utf-8')

    # 获取md5实例
    md5 = hashlib.md5()

    # 拿取uuid的md5摘要
    md5.update(uuid_str)

    # 返回固定长度的字符串
    return md5.hexdigest()

def upload_img(request):
    """接收上传的文件"""
    # 接收上传的文件
    rev_file = request.FILES.get('avatar')
    # 判断，是否有文件
    if not rev_file:
        return JsonResponse({'code':0, 'msg':'图片不存在！'})
    # 获得一个唯一的名字： uuid +hash
    new_name = get_unique_id()
    # 准备写入的URL
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(rev_file.name)[1] )
    # 开始写入到本次磁盘
    try:
        f = open(file_path,'wb')
        # 多次写入
        for i in rev_file.chunks():
            f.write(i)
        # 要关闭
        f.close()
        # 返回
        return JsonResponse({'code': 1, 'name': new_name + os.path.splitext(rev_file.name)[1]})

    except Exception as e:
        return JsonResponse({'code':0, 'msg':str(e)})

def read_excel_dict(path: str):
    """读取Excel的数据，存储为字典---[{},{},{}...]"""
    # 实例化一个workbook
    workbook = openpyxl.load_workbook(path)
    # 实例化一个sheet
    sheet = workbook['students']
    # 定义一个变量存储最终的数据--[]
    students = []
    # 准备key
    keys = ['sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address', 'image']
    # 遍历
    for row in sheet.rows:
        # 定义一个临时的字典
        temp_dict = dict()
        # 组合值和key
        for index, cell in enumerate(row):
            # 组合
            temp_dict[keys[index]] = cell.value
        # 附件到list中
        students.append(temp_dict)
    return students

def import_students_excel(request):
    """从Excel批量导入学生学生信息"""
    # =====1.接收Excel文件存储到Media文件夹
    receive_file = request.FILES.get('excel')
    if not receive_file:
        return JsonResponse({'code':0, 'msg':'Excel文件不存在！'})

    try:
        # 这里我的实现是文件名都统一改为excel_for_students。
        # 这样后面再上传文件的时候就直接把前面的覆盖了，不然以后会有特别多的文件。
        new_name = 'excel_for_students'
        file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(receive_file.name)[1])
        f = open(file_path, 'wb')
        # 多次写入
        for i in receive_file.chunks():
            f.write(i)
        # 要关闭
        f.close()

    except Exception as e:
        return JsonResponse({'code':0, 'msg':str(e)})

    # =====2.读取存储在Media文件夹的数据
    students = read_excel_dict(file_path)
    # =====3.把读取的数据存储到数据库
    # 定义几个变量：success，error， errors
    success, error, errors = 0, 0, []

    # 开始遍历
    for stu in students:
        try:
            # keys = ['sno', 'name', 'gender', 'birthday', 'mobile', 'email', 'address', 'image']
            # 如果找到了，说明这个学号已经存在。
            if Student.objects.filter(Q(sno__icontains=stu['sno'])):
                error += 1
                errors.append(stu['sno'])

            else:
                obj_student = Student(sno=stu['sno'], name=stu['name'], gender=stu['gender'],
                                      birthday=stu['birthday'], mobile=stu['mobile'],
                                      email=stu['email'], address=stu['address'],
                                      image='041abd5bcf337655efb2443e5ae57a94.jpg')

                obj_student.save()
                # 计数
                success += 1

        except:
            # 如果失败了
            error += 1
            errors.append(stu['sno'])

    # 4.返回--导入信息（成功：5， 失败4-- （sno）），所有学生
    obj_students = Student.objects.all().values()
    students = list(obj_students)
    return JsonResponse({'code': 1, 'success': success, 'error': error, 'errors': errors, 'data': students})


def write_to_excel(data:list, path:str):
    """把数据库写入到Excel"""
    # 实例化一个workbook
    workbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = workbook.active
    # 为sheet命名
    sheet.title = 'students'
    # 准备keys
    keys = data[0].keys()
    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每一个元素
        for k, v in enumerate(keys):
            sheet.cell(row=index + 1, column=k + 1, value=str(item[v]))
    # 写入到文件
    workbook.save(path)


def export_students(request):
    """导出数据到Excel"""

    try:
        # 获取所有的学生信息
        obj_students = Student.objects.all().values()
        # 转为list
        students = list(obj_students)
        # 准备写入的路径
        path = os.path.join(settings.MEDIA_ROOT, 'students_to_excel.xlsx')
        # 写入到excel
        write_to_excel(students, path)

        # 返回
        return JsonResponse({'code':1, 'data': 'students_to_excel.xlsx'})
    except Exception as e:
        return JsonResponse({'code':0, 'msg':str(e)})